def _get_namedrange(book, rangename, sheetname=None):
    """Get range from a workbook.

    A workbook can contain multiple definitions for a single name,
    as a name can be defined for the entire book or for
    a particular sheet.

    If sheet is None, the book-wide def is searched,
    otherwise sheet-local def is looked up.

    Args:
        book: An openpyxl workbook object.
        rangename (str): Range expression, such as "A1", "$G4:$K10",
            named range "NamedRange1".
        sheetname (str, optional): None for book-wide name def,
            sheet name for sheet-local named range.

    Returns:
        Range object specified by the name.

    """

    def cond(namedef):

        if namedef.type.upper() == "RANGE":
            if namedef.name.upper() == rangename.upper():

                if sheetname is None:
                    if not namedef.localSheetId:
                        return True

                else:  # sheet local name
                    sheet_id = [sht.upper() for sht in book.sheetnames].index(
                        sheetname.upper()
                    )

                    if namedef.localSheetId == sheet_id:
                        return True

        return False

    def get_destinations(name_def):
        """Workaround for the bug in DefinedName.destinations"""

        from openpyxl.formula import Tokenizer
        from openpyxl.utils.cell import SHEETRANGE_RE

        if name_def.type == "RANGE":
            tok = Tokenizer("=" + name_def.value)
            for part in tok.items:
                if part.subtype == "RANGE":
                    m = SHEETRANGE_RE.match(part.value)
                    if m.group("quoted"):
                        sheet_name = m.group("quoted")
                    else:
                        sheet_name = m.group("notquoted")

                    yield sheet_name, m.group("cells")

    namedef = next(
        (item for item in book.defined_names.definedName if cond(item)), None
    )

    if namedef is None:
        return None

    dests = get_destinations(namedef)
    xlranges = []

    sheetnames_upper = [name.upper() for name in book.sheetnames]

    for sht, addr in dests:
        if sheetname:
            sht = sheetname
        index = sheetnames_upper.index(sht.upper())
        xlranges.append(book.worksheets[index][addr])

    if len(xlranges) == 1:
        return xlranges[0]
    else:
        return xlranges